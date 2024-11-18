from fastapi import Depends, HTTPException, APIRouter, Body
from jwt_manager import JWTBearer
import base64
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill

# Definir el enrutador para la API de archivo base64
archivo_base64_router = APIRouter()


@archivo_base64_router.post(
    "/api2/ArchivoBase64/", tags=['Archivos'], dependencies=[Depends(JWTBearer())]
)
async def generar_base64_desde_archivo(
    ruta_archivo: str = Body(..., embed=True, description="Ruta al archivo")
) -> dict:
    """
    Genera una representación base64 de un archivo localizado en la ruta especificada.

    Argumentos:
    ruta_archivo (str): La ruta del archivo que se quiere convertir en base64.

    Retorna:
    dict: Un diccionario con el contenido base64 del archivo, su nombre y tipo.
    """
    if not os.path.exists(ruta_archivo):
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    # Abrir el archivo Excel y formatearlo
    workbook = load_workbook(ruta_archivo)
    sheet = workbook.active

    # Crear una tabla con los datos en la hoja activa
    min_row, max_row = sheet.min_row, sheet.max_row
    min_col, max_col = sheet.min_column, sheet.max_column
    table_ref = f"{sheet.cell(row=min_row, column=min_col).coordinate}:" \
                f"{sheet.cell(row=max_row, column=max_col).coordinate}"
    tabla = Table(displayName="TablaDatos", ref=table_ref)

    # Estilo de la tabla
    estilo_tabla = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tabla.tableStyleInfo = estilo_tabla
    sheet.add_table(tabla)

    # Ajustar el ancho de las columnas según el contenido
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        col_letter = col[0].column_letter
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Aplicar color a las celdas con el valor "SI"
    fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="lightGrid")
    for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value == "SI":
                cell.fill = fill_color

    # Guardar el archivo modificado en un archivo temporal
    temp_file_path = "temp_files/temp_formatted.xlsx"
    workbook.save(temp_file_path)

    # Convertir el archivo a base64
    with open(temp_file_path, 'rb') as archivo:
        contenido_binario = archivo.read()
        contenido_base64 = base64.b64encode(contenido_binario).decode('utf-8')

    # Eliminar el archivo temporal después de la conversión
    os.remove(temp_file_path)

    nombre_archivo = os.path.basename(ruta_archivo)
    tipo_archivo = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if nombre_archivo.endswith(".xlsx") else "application/octet-stream"
    )

    respuesta_archivo = {
        "content": contenido_base64,
        "name": nombre_archivo,
        "type": tipo_archivo
    }

    return respuesta_archivo




